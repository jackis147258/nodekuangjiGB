import asyncio
from django.contrib import admin

# Register your models here.

from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from .models import CustomUser,ebcJiaSuShouYiJiLu ,tokenZhiYaJiShi,payToken,userToken

from django.http import HttpResponse, JsonResponse

from simpleui.admin import AjaxAdmin


from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from app1.tasks import tokenNewUserList
from apiV1.DeFi2 import  getTokenNewUser
from .ebcFenRun import FenRun,createEbcUser
from .ebcUserTiXian import  userTiXian ,userTiXianHash

from django.urls import reverse
from django.utils.html import format_html
from . import  setTokenUsersBalance,web3_utils,web3_tixian
from django.db import transaction



from django.contrib.auth import get_user_model
User = get_user_model()
from typing import Optional

from .nodeKjFenRun import fanTiXianTime
# import web3_utils 
def export_excel(modeladmin, request, queryset):
    # 创建一个 Excel 工作簿
    wb = Workbook()
    ws = wb.active

    # 添加标题行（字段名）
    # field_names = [field.name for field in modeladmin.model._meta.fields]
    # 添加标题行（字段名），并过滤掉 'uid' 字段
    field_names = [field.name for field in modeladmin.model._meta.fields if field.name != 'uid']

    for col_num, column_title in enumerate(field_names, 1):
        col_letter = chr((col_num - 1) % 26 + ord('A'))
        ws['{}1'.format(col_letter)] = column_title

    # 添加数据行
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(field_names, 1):
            col_letter = chr((col_num - 1) % 26 + ord('A'))
            
            value = getattr(obj, field_name)
            # 如果值是带有时区的datetime对象，则移除时区信息
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            
            # ws['{}{}'.format(col_letter, row_num)] = getattr(obj, field_name)
            ws['{}{}'.format(col_letter, row_num)] = value


    # 创建 HTTP 响应对象
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_exported_data.xlsx'
    
    # 将工作簿保存到响应对象
    wb.save(response)
    return response

export_excel.short_description = "导出到Excel"

@admin.register(CustomUser)
class CustomUserAdmin(AjaxAdmin):
  
    
     
    list_display = ("id","username", "statusTiXian","userStakesA","userStakesB", "userStakesBfanHuan", "fanHuan","EbcCreated_at", "EbcLastFanHuan_at", "status",
                    "parent",  )
    actions = ('layerGetTokenUserAll','layerCreateToken',export_excel)
    list_filter = ('status',  )  # 添加筛选器
    # search_fields = ('username', 'id', 'userStakesA', 'userStakesB', 'userStakesBfanHuan', 'parent', 'EbcCreated_at', 'EbcLastFanHuan_at', 'fanHuan')  # 添加搜索字段
    search_fields = ('username', 'id',  )  # 添加搜索字段
    list_per_page = 300  # 设置每页显示的条目数

     # 添加按钮 显示 和关闭
    # def has_add_permission(self, request):
    #     return False
    # def has_delete_permission(self, request, obj=None):
    #     return False
    
    # class Media:
    #     js = ('js/your_script.js',)
    #     css = {
    #         'all': ('admin_styles.css',)  # 更新为您CSS文件的实际路径
    #     } 
    
    # def custom_button(self, obj):
    #     url = reverse('viewsEbc:changFuLeiview')
    #     return format_html('<a class="button" href="{}">Custom Action</a>', url)

    # custom_button.short_description = 'changF'

        

    
        # 获取链上用户
    def layerGetTokenUserAll(self, request, queryset):
        
        # reMsg=self.workClose(request, queryset)
        print('sssssss')
        reMsg='建立成功'        
        
        acct=request.user
        post = request.POST
        name=post.get('name')        
        try:
            int_value = int(name)    
            
            if int_value==3:
                t_return= FenRun()                      
            # 处理用户 设置 余额
            if int_value==2: 
                setTokenUsersBalance.userBlance();
            # asyncio.run(main())
            
            if int_value==1:
                t_return= getTokenNewUser.work()
                if t_return is not None:
                    createEbcUser(t_return)
            # t_taskId = tokenNewUserList.delay()
            
            # print("Conversion successful:", int_value)
        
        except Exception as e:  
            result = ["Failed -defi.runCode", f"ERROR: {e}" ]
            print(result) 
            # return 'err'
            return JsonResponse(data={
                    'status': 'err',
                    'msg':  str(e)
                })
        except ValueError:
            # print("Conversion failed: The value is not convertible to int.")
            return JsonResponse(data={
                'status': 'success',
                'msg':  "请输入数字"
            })
      
        return JsonResponse(data={
                'status': 'success',
                'msg': reMsg
            })


    layerGetTokenUserAll.short_description = '1:获取链上用户 3.分润 fenrun 2:设置用户数据可取额度'
    layerGetTokenUserAll.type = 'success'
    layerGetTokenUserAll.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    layerGetTokenUserAll.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        # 'title': '弹出层输入框',
        # 提示信息
        'tips': '获取链上用户',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [
            {
            # 这里的type 对应el-input的原生input属性，默认为input
            'type': 'input',
            # key 对应post参数中的key
            'key': 'name',
            # 显示的文本
            'label': '开始数字',
            # 为空校验，默认为False
            'require': True
        },       
                   ]
    }
    
    
    # 生成钱包
    def layerCreateToken(self, request, queryset):
        
        # reMsg=self.workClose(request, queryset)
        print('sssssss')
        reMsg='建立成功'        
        
        acct=request.user
        post = request.POST
        name=post.get('name')
        
        try:
            int_value = int(name)
            # print("Conversion successful:", int_value)
        except ValueError:
            # print("Conversion failed: The value is not convertible to int.")
            return JsonResponse(data={
                'status': 'success',
                'msg':  "请输入数字"
            }) 
        return JsonResponse(data={
                'status': 'success',
                'msg': reMsg
            })


    layerCreateToken.short_description = '生成钱包'
    layerCreateToken.type = 'success'
    layerCreateToken.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    layerCreateToken.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        # 'title': '弹出层输入框',
        # 提示信息
        'tips': '生成钱包',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [
            {
            # 这里的type 对应el-input的原生input属性，默认为input
            'type': 'input',
            # key 对应post参数中的key
            'key': 'name',
            # 显示的文本
            'label': '生成数量',
            # 为空校验，默认为False
            'require': True
        },       
                   ]
    }
    


@admin.register(ebcJiaSuShouYiJiLu)
class  ebcJiaSuShouYiJiLu(AjaxAdmin):
    
    fieldsets = (
        (None, {
            'fields':  ('uidA','uidB','status' ,'fanHuan','Layer','liuShuiId','Remark' , )
        }),
    )


    # add_fieldsets = UserAdmin.add_fieldsets + (
    #     (None, {'fields': ('parent','email')}),
    # )
    
        # actions = ('layer_input','layer_close',perform_check_all)
    list_display = ('id','liuShuiId','uidA','uidB','status' ,'created_at','Layer','fanHuan','Remark' ,       )
    
    list_filter = ('status', 'Layer')  # 添加筛选器
    search_fields = ('uidB', 'id')  # 添加搜索字段
        
    # actions = [export_excel ]
    
    # actions = ('layerGetTokenUserAll','layerCreateToken',export_excel)
    
    
  
@admin.register(tokenZhiYaJiShi)
class  tokenZhiYaJiShiAdmin(AjaxAdmin):
    
    # fieldsets = (
    #     (None, {
    #         'fields':  ('tokenName','number','zhiYaTime' ,'kaiShiTime','status','uid' ,'Remark', )
    #     }),
    # ) 
    
        # actions = ('layer_input','layer_close',perform_check_all)
    list_display = ('id','tokenName','number','zhiYaTime' ,'kaiShiTime','status','uid' ,'Remark', 'statusTiXian', )
    
    list_filter = ('status', 'zhiYaTime')  # 添加筛选器
    search_fields = ('tokenName', 'uid__username', 'statusTiXian',)  # 添加搜索字段


# 菜单 提现支付流水
@admin.register(payToken)
class  payTokenAdmin(AjaxAdmin):
    
     # 添加按钮 显示 和关闭
    def has_add_permission(self, request):
        return False
    # def has_delete_permission(self, request, obj=None):
    #     return False
    
    fieldsets = (
        (None, {
            
            'fields':  ('TxHash','uidB','status' ,'amount','Layer','liuShuiId','Remark','HashId' , )
        }),
    )


    # add_fieldsets = UserAdmin.add_fieldsets + (
    #     (None, {'fields': ('parent','email')}),
    # )
    
    # actions = ('layer_input','layer_close',perform_check_all)
    # actions = ('tiXianFanHuan',)
    actions = ('layerTokenTiXianUserAll','tiXianFanHuan',export_excel)



    list_display = ('id','liuShuiId','TxHash','uidB','status' ,'created_at','Layer','amount','Remark' , 'fanTiXian'  )
    
    list_filter = ('status', 'Layer')  # 添加筛选器
    search_fields = ('uidB', 'id')  # 添加搜索字段

    def tiXianFanHuan(self, request, queryset): 
        reMsg='导入成功:'                
        acct=request.user 
        data_text = request.POST.get('data')   
        # 使用数据库事务来确保所有操作都成功完成
        fanTiXianTime(data_text)
             
        # return HttpResponseRedirect('/admin/app1/t_tokenaddr/')  # 重定向到列表页面
      
        return JsonResponse(data={
                'status': 'success',
                'msg': reMsg
            }) 
   
    tiXianFanHuan.short_description = '提现返还'
    tiXianFanHuan.type = 'success'
    tiXianFanHuan.icon = 'el-icon-s-promotion'     
    tiXianFanHuan.layer = {
    'tips': '提现返还',
    'confirm_button': '确认提现返还',
    'cancel_button': '取消',
    'width': '50%',
    'labelWidth': "100px",
    'params': [
       
        {
            'type': 'textarea',
            'key': 'data',
            'label': '数据',
            'require': True,
            'placeholder': '每行一个数据'
        }
    ]
}
    

    # 每行按钮    
    def fanTiXian(self, obj):
        # CUS2 = models.CommonUserStation2.objects.get(id=obj.tableid)
        # parameter_str = 'id={}&divpercent={}'.format(str(obj.id), str(CUS2.divpercent))
        parameter_str = 'id={} '.format(str(obj.id) )
        btn_str = '<a class="btn btn-xs btn-danger" href="{}">' \
                    '<input name=" 返提现款"' \
                    'type="button" id="passButton" ' \
                    'title="passButton" value="返提现款">' \
                    '</a>'
        return format_html(btn_str, '/reg/fanTiXian/?{}'.format(parameter_str))

    fanTiXian.short_description = '返提现款'  
        
     
    
        # 执行提现
    def layerTokenTiXianUserAll(self, request, queryset):
        
        # reMsg=self.workClose(request, queryset)
        print('sssssss')
        reMsg='建立成功'        
        
        acct=request.user
        post = request.POST
        name=post.get('name')        
        try:
            int_value = int(name)      
            # asyncio.run(main())
            if int_value==1:
                t_return= userTiXian()
            
            if int_value==2:
                t_return=userTiXianHash()


            if int_value==3:
                web3_utils.listen_to_deposit_events()
            # 提现
            if int_value==4:
                web3_tixian.listen_to_Withdrawal_events()

              # 充值 单独
            if int_value==5:                
                qukuai=post.get('qukuai')  
                web3_utils.listenDepositOne(qukuai)
            

                # 提现 单独
            if int_value==6:                
                qukuai=post.get('qukuai')  
                web3_tixian.listen_to_Withdrawal_eventsOne(qukuai)
            
            # if t_return is not None:
            #     createEbcUser(t_return)
            # t_taskId = tokenNewUserList.delay()
            
            # print("Conversion successful:", int_value)
        
        except Exception as e:  
            result = ["Failed -defi.runCode", f"ERROR: {e}" ]
            print(result) 
            # return 'err'
            return JsonResponse(data={
                    'status': 'err',
                    'msg':  str(e)
                })
        except ValueError:
            # print("Conversion failed: The value is not convertible to int.")
            return JsonResponse(data={
                'status': 'success',
                'msg':  "请输入数字"
            })
      
        return JsonResponse(data={
                'status': 'success',
                'msg': reMsg
            })


    layerTokenTiXianUserAll.short_description = '提现'
    layerTokenTiXianUserAll.type = 'success'
    layerTokenTiXianUserAll.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    layerTokenTiXianUserAll.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        # 'title': '弹出层输入框',
        # 提示信息
        'tips': '提现处理',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [
            {
            # 这里的type 对应el-input的原生input属性，默认为input
            'type': 'input',
            # key 对应post参数中的key
            'key': 'name',
            # 显示的文本
            'label': '1提现 2处理hash 3充值 4 提现 5 充值区块 6提现区块' ,
            # 为空校验，默认为False
            'require': True
        },      
         {
            # 这里的type 对应el-input的原生input属性，默认为input
            'type': 'input',
            # key 对应post参数中的key
            'key': 'qukuai',
            # 显示的文本
            'label': '区块数字',    
             # 设置默认值
            'value': '3546455'    ,    
            # 为空校验，默认为False
            'require': True
        }, 
                   ]
    }


    

@admin.register(userToken)
class  userTokenAdmin(AjaxAdmin):
    
    # fieldsets = (
    #     (None, {
    #         'fields':  ('uidA','uidB','status' ,'fanHuan','Layer','liuShuiId','Remark' , )
    #     }),
    # )


    # add_fieldsets = UserAdmin.add_fieldsets + (
    #     (None, {'fields': ('parent','email')}),
    # )
    
        # actions = ('layer_input','layer_close',perform_check_all)
    list_display = ('id','usdtToken','ylToken','jzToken','status' ,'uid','cTime' ,'Remark' ,       )
    
    list_filter = ('status', 'usdtToken')  # 添加筛选器
    # search_fields = ('uid', 'id')  # 添加搜索字段
    search_fields = ('id', 'uid__username')  # 添加搜索字段

        
    