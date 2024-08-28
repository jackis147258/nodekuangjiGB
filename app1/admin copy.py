from django.contrib import admin
from app1.myFroms import  infoFrom,TokenForm,CategoryForm
from django.utils.html import format_html
from apiV1.DeFi2 import DeFi2
from multiprocessing import Process
from concurrent.futures import ThreadPoolExecutor, as_completed

from app1.until.MyThreadPoolExecutor import global_thread_pool 

from app1.tasks import taskQuantify
from celery.app.control import Control
from moon39.celery   import app
from app1.models import T_task,T_TokenAddr,T_TokenAddrBak,CategoryToken  #保存 task id 以后删除
import datetime
from apiV1.DeFi2 import  createWallet , GetBnbWeb
from cryptography.fernet import Fernet
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect,reverse,render
from django.db import transaction
from django.http import HttpResponseRedirect




# from celery.task.control import revoke



# from demoapp.tasks import add, mul, xsum
# Register your models here.
# admin.site.register(News)
# admin.site.register(NewsType)
admin.site.site_header = "姨妈量化"
admin.site.index_title = "后台"



# from .models import T_task,T_trade  , TcSport,TcOneTwo,dengLuInfo,webInfo,T_Quantify1,T_TokenAddr
from .models import T_task,T_trade,webInfo,T_Quantify1,T_TokenAddr
# admin.site.register(TcWeb) 

future =None 
class GroupAdmin(admin.ModelAdmin):
    form = infoFrom


# from import_export.admin import ExportActionModelAdmin
from django.http import HttpResponse, JsonResponse
# from simplepro.dialog import ModalDialog
from simpleui.admin import AjaxAdmin

# def perform_check_all(modeladmin, request, queryset):
#     print("ok")
#     for obj in queryset:
#         # adapayment = AdapayPayment.objects.get(id=int(obj.id))
#         adapayment =  int(obj.id) 
#         adapayment.closeexpend = '1'
#         adapayment.save() 
# perform_check_all.short_description = '执行对账'

def convert_integer_str_to_18_decimal_places(integer_str):
    integer_str = integer_str[::-1]
    while len(integer_str) < 18:
        integer_str += "0"
    integer_str = integer_str[::-1]
    return f"0.{integer_str}"



def TokenBuySellZero(modeladmin, request, queryset):
    print("ok")
    # T_TokenAddrList=T_TokenAddr.objects.filter(uid=acct.id,status=1)     
    for obj in queryset:  
        obj.BuyTOKEN_price = ''
        obj.SellTOKEN_price = ''
        obj.SellNumber = 0      
        obj.BuyNumber = 0
        obj.save()

    # for obj in queryset:
    #     # adapayment = AdapayPayment.objects.get(id=int(obj.id))
    #     adapayment =  int(obj.id) 
    #     adapayment.closeexpend = '1'
    #     adapayment.save() 
TokenBuySellZero.short_description = '买卖清零'
TokenBuySellZero.type = 'success'
TokenBuySellZero.icon = 'el-icon-s-promotion'


def allTokenPrice(modeladmin, request, queryset):
    print("ok")
    # T_TokenAddrList=T_TokenAddr.objects.filter(uid=acct.id,status=1) 
    token_private_list = [obj  for obj in queryset]

    ImputationList=GetBnbWeb.GetBnbPrice(request,token_private_list)   
    print (str(ImputationList))

    for item in ImputationList:
        token_addr = item[1].split(":")[1]  # 分割字符串取得token地址
        price = item[3]

        price=convert_integer_str_to_18_decimal_places(str(price))

        print(convert_integer_str_to_18_decimal_places(price))  # 输出：0.000000000000001234


        # 查找并更新
        token_obj = T_TokenAddr.objects.filter(TOKEN_ADDRESS=token_addr).first()
        if token_obj:
            token_obj.TOKEN_price = str(price)
            token_obj.save()

    # for obj in queryset:
    #     # adapayment = AdapayPayment.objects.get(id=int(obj.id))
    #     adapayment =  int(obj.id) 
    #     adapayment.closeexpend = '1'
    #     adapayment.save() 
allTokenPrice.short_description = 'bnb余额'
allTokenPrice.type = 'success'
allTokenPrice.icon = 'el-icon-s-promotion'


 


from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime

def export_excel(modeladmin, request, queryset):
    # 创建一个 Excel 工作簿
    wb = Workbook()
    ws = wb.active

    # 添加标题行（字段名）
    # field_names = [field.name for field in modeladmin.model._meta.fields]
    # 添加标题行（字段名），并过滤掉 'uid' 字段
    field_names = [field.name for field in modeladmin.model._meta.fields if field.name != 'uid']

    for col_num, column_title in enumerate(field_names, 1):
        col_letter = chr((col_num - 1) % 26 + ord('A'))
        ws['{}1'.format(col_letter)] = column_title

    # 添加数据行
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field_name in enumerate(field_names, 1):
            col_letter = chr((col_num - 1) % 26 + ord('A'))
            
            value = getattr(obj, field_name)
            # 如果值是带有时区的datetime对象，则移除时区信息
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            
            # ws['{}{}'.format(col_letter, row_num)] = getattr(obj, field_name)
            ws['{}{}'.format(col_letter, row_num)] = value


    # 创建 HTTP 响应对象
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=my_exported_data.xlsx'
    
    # 将工作簿保存到响应对象
    wb.save(response)
    return response

export_excel.short_description = "导出到Excel"


@admin.register(T_Quantify1) 
class T_Quantify1Admin(AjaxAdmin):
    
     #    列表部分
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        accUser = request.user 
        # 在此处添加任何过滤逻辑
        queryset = queryset.filter(uid=accUser)
        return queryset
    
    def save_model(self, request, obj, form, change):
        if not change:
            # 如果是创建新的对象，为 uid 字段设置当前登录的用户
            obj.uid = request.user
        super().save_model(request, obj, form, change)
    #    列表部分  'status',
    
    # list_display = ('TRADE_TOKEN_ADDRESS','buyPrice', 'sellPrice','taskLog', 'get_run','taskClose',
    #                   'USDT_TOKEN_ADDRESS', 'WBNB_ADDRESS', 'LP_ymiiUsdt_ADDRESS', 'Tools_lpPrice_ADDRESS', 
    #                  'PANCAKE_ROUTER_ADDRESS', 'Bnb_TokenGas', 'buyNumberTotal',  'coinsRecharge', 'AmountBuy0', 'AmountBuy1',
    #                  'AmountSell0', 'AmountSell1', 'Token_Transaction_Time', 'PercentBuy','PercenttSell',
    #                 )
    
    list_display = ('TRADE_TOKEN_ADDRESS','id','buyPrice', 'sellPrice','taskLog', 'get_run','taskClose',
                       'coinsRecharge', 'AmountBuy0', 'AmountBuy1',
                    'AmountSell0', 'AmountSell1', 'Token_Transaction_Time', 'start_time', 'end_time',
                     'duration',
                )
    

    
    list_editable = ('buyPrice','sellPrice',)
    
    date_hierarchy = 'created_at'
    
    list_per_page = 15
    
    # actions = ('layer_input','layer_close',perform_check_all)
    
    actions = [export_excel ]
     
    
       
    def taskLog(self, obj):
        
        parameter_str = '{}'.format(str(obj.id) )
        btn_str = '<a class="btn btn-xs btn-danger" target="_blank" href="{}">' \
                  '<input name="查看明细"' \
                  'type="button" id="passButton" ' \
                  'title="passButton" value="查看明细">' \
                  '</a>'
        return format_html(btn_str, '/app1/showlog/{}/'.format(parameter_str))
    taskLog.short_description = '明细'

    def get_run(self, obj):
        
        parameter_str = 'id={} '.format(str(obj.id) )
        btn_str = '<a class="btn btn-xs btn-danger" href="{}">' \
                  '<input name="运行"' \
                  'type="button" id="passButton" ' \
                  'title="passButton" value="运行">' \
                  '</a>'
        return format_html(btn_str, '/app1/QuantifyRun/?{}'.format(parameter_str))

    get_run.short_description = '运行'
    
    
    def taskClose(self, obj):
       
        # parameter_str = '{}'.format(str(obj.id) )   target="_blank" 
        btn_str = '<a class="btn btn-xs btn-danger"    href="{}">' \
                    '<input name="查看运行"' \
                    'type="button" id="passButton" ' \
                    'title="passButton" value="查看运行">' \
                    '</a>'
                    
                    
        return format_html(btn_str, '/admin/app1/t_task/' )

    taskClose.short_description = '查看运行' 
    
    # fieldsets = [('价格调整',{'fields':[('buyPrice','sellPrice'),
    #                  'coinsRecharge', 'AmountBuy0', 'AmountBuy1',
    #                 'AmountSell0', 'AmountSell1', 'Token_Transaction_Time', 'PercentBuy','PercenttSell', 'buyNumberTotal',
    #                             ]}),
    #              (u'Token设置',{'fields':['Bnb_TokenGas', 
    #                  'TRADE_TOKEN_ADDRESS', 'USDT_TOKEN_ADDRESS', 'WBNB_ADDRESS', 'LP_ymiiUsdt_ADDRESS', 'Tools_lpPrice_ADDRESS', 
    #                  'PANCAKE_ROUTER_ADDRESS', 
    #                 'status',  'created_at'
    #                  ]})]
    
    
    fieldsets = [('价格调整',{'fields':[('buyPrice','sellPrice'),
                     'coinsRecharge', 'AmountBuy0', 'AmountBuy1',
                    'AmountSell0', 'AmountSell1', 'Token_Transaction_Time', 
                                ]}),
                (u'单独账户',{'fields':['Dlprice_status','PercenttSell','DlBuyNumber','DlSellNumber',   
                     ]}),
                (u'设置开始时间',{'fields':['start_time',  
                    ]}),
                # 'end_time','duration',  
                     
                (u'Token设置',{'fields':['Bnb_TokenGas', 
                     'TRADE_TOKEN_ADDRESS','LP_ymiiUsdt_ADDRESS',   
                     ]})]
   
    readonly_fields = ( 'created_at',) 
    def layer_input(self, request, queryset):
     
        # 根据 id 建立进程运行。
        self.work(request, queryset)
      
        post = request.POST
        # 这里获取到数据后，可以做些业务处理
        # post中的_action 是方法名
        # post中 _selected 是选中的数据，逗号分割
        if not post.get('_selected'):
            return JsonResponse(data={
                'status': 'error',
                'msg': '请先选中数据！'
            })
        else:
            return JsonResponse(data={
                'status': 'success',
                'msg': '处理成功！'
            })

    layer_input.short_description = '运行量化'
    layer_input.type = 'success'
    layer_input.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    layer_input.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        'title': '弹出层输入框',
        # 提示信息
        'tips': ' 开始执行选中量化程序',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        # 'params': [{
        #     # 这里的type 对应el-input的原生input属性，默认为input
        #     'type': 'input',
        #     # key 对应post参数中的key
        #     'key': 'name',
        #     # 显示的文本
        #     'label': '名称',
        #     # 为空校验，默认为False
        #     'require': True
        # },   ]
    }

    def layer_close(self, request, queryset):
        
        reMsg=self.workClose(request, queryset)
        
        global future
        
        # celery_control = Control(app=app) 
        # celery_control.revoke(str(future.id), terminate=True)   
        
      
        return JsonResponse(data={
                'status': 'success',
                'msg': reMsg
            })


    layer_close.short_description = '关闭量化'
    layer_close.type = 'success'
    layer_close.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    layer_close.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        'title': '弹出层输入框',
        # 提示信息
        'tips': '关闭选中量化程序',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        # 'params': [{
        #     # 这里的type 对应el-input的原生input属性，默认为input
        #     'type': 'input',
        #     # key 对应post参数中的key
        #     'key': 'name',
        #     # 显示的文本
        #     'label': '名称',
        #     # 为空校验，默认为False
        #     'require': True
        # },   ]
    }

    
    def work(self,request, queryset): 
        # global future
		# 判断worker线程是否正在运行，没有则唤醒
  
        acct=request.user
        for obj in queryset:            
            # taskId = taskQuantify.delay(obj.id)
            t_objId=obj.id
            t_taskId = taskQuantify.delay(t_objId)
 
            
            t_task=T_task ()
            t_task.taskId=str(t_taskId.id)
            t_task.QuantifyId=obj.id
            t_task.taskRemark='remark'
            t_task.uid=acct
            t_task.save()
                    
 
 

    
    def workClose(self,request, queryset): 
        global future
        celery_control = Control(app=app)   
        
        time1 = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')                
        # result = [time1,'出售地址:',self.walletAddress ,f"出售完成: {self.web3.fromWei(tokenToSell, self.TradingTokenDecimal)}  " ,'hash:',self.web3.to_hex(tx_token),]
        result = [time1,'已关闭:']

        for obj in queryset:     
            t_task=T_task.objects.filter(QuantifyId=obj.id,status=1).order_by('-id').first()
            if not t_task==None:
                reCelery=celery_control.revoke(str(t_task.taskId), terminate=True)  
            else:
                result = [time1,'没有找到 taskId']
        return result
                
 




@admin.register(T_task) 
class T_taskAdmin(AjaxAdmin):  
   
    #    列表部分
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        accUser = request.user 
        # 在此处添加任何过滤逻辑
        queryset = queryset.filter(uid=accUser,status=1)
        return queryset
    
    def has_add_permission(self, request):
        return False
    def has_delete_permission(self, request, obj=None):
        return False
    
    list_display = ('taskId', 'status', 'QuantifyId', 'created_at' ,'taskLog', 'taskStop',) 
 
    date_hierarchy = 'created_at'
    
    list_per_page = 15

      # form 部分
    fields =('taskId',   'status','QuantifyId',)

    readonly_fields = ('status', 'taskId','QuantifyId','created_at',)
    
    # actions = [perform_check_all ]
    
    
    # /app1/chat/' + roomName + '/' 
    
    def taskLog(self, obj):
        # CUS2 = models.CommonUserStation2.objects.get(id=obj.tableid)
        # parameter_str = 'id={}&divpercent={}'.format(str(obj.id), str(CUS2.divpercent))
        parameter_str = '{}'.format(str(obj.QuantifyId) )
        btn_str = '<a class="btn btn-xs btn-danger" href="{}">' \
                  '<input name="查看明细"' \
                  'type="button" id="passButton" ' \
                  'title="passButton" value="查看明细">' \
                  '</a>'
                  
                  
        return format_html(btn_str, '/app1/showlog/{}/'.format(parameter_str))

    taskLog.short_description = '明细'


    def taskStop(self, obj):
        # CUS2 = models.CommonUserStation2.objects.get(id=obj.tableid)
        # parameter_str = 'id={}&divpercent={}'.format(str(obj.id), str(CUS2.divpercent))
        parameter_str = 'id={} '.format(str(obj.id) )
        btn_str = '<a class="btn btn-xs btn-danger" href="{}">' \
                    '<input name="停止"' \
                    'type="button" id="passButton" ' \
                    'title="passButton" value="停止">' \
                    '</a>'
        return format_html(btn_str, '/app1/stopTask/?{}'.format(parameter_str))

    taskStop.short_description = '停止' 

from django.contrib.admin import SimpleListFilter


class UserCategoryFilter(SimpleListFilter):
    title = '地址分类'  # Human readable title, used in the right admin sidebar
    parameter_name = '地址分类'  # URL parameter that will be used when the filter is selected.

    def lookups(self, request, model_admin):
        # Only return categories owned by the logged in user
        return [(cat.id, cat.name) for cat in CategoryToken.objects.filter(uid=request.user)]

    def queryset(self, request, queryset):
        # If a specific category is chosen, filter by that category.
        if self.value():
            return queryset.filter(category__id=self.value())
        return queryset

@admin.register(T_TokenAddr) 
class T_TokenAddrAdmin(AjaxAdmin):
    form = TokenForm
    

    def get_form(self, request, obj=None, **kwargs):
        Form = super(T_TokenAddrAdmin, self).get_form(request, obj, **kwargs)
        
        class RequestForm(Form):
            def __new__(cls, *args, **kwargs):
                kwargs['user'] = request.user
                return Form(*args, **kwargs)
        
        return RequestForm
    
    def save_model(self, request, obj, form, change):
        if not change:
            # 如果是创建新的对象，为 uid 字段设置当前登录的用户
            obj.uid = request.user
        super().save_model(request, obj, form, change)
    # 添加按钮 显示 和关闭
    def has_add_permission(self, request):
        return True
    # def has_delete_permission(self, request, obj=None):
    #     return False
    
    class Media:
        js = ('js/your_script.js',)
        css = {
            'all': ('admin_styles.css',)  # 更新为您CSS文件的实际路径
        }

   

    #    列表部分
    def get_queryset(self, request):

        queryset = super().get_queryset(request)

        accUser = request.user 

        # 在此处添加任何过滤逻辑
        queryset = queryset.filter(uid=accUser)
        # queryset = queryset.filter(uid=accUser,status=1)
        return queryset
    
    list_display = ('TOKEN_ADDRESS', 'TOKEN_price', 'TOKEN_otherPrice', 'TOKEN_bak', 'status',                    
                    'BuyTOKEN_price', 'SellTOKEN_price',  'SellNumber', 'BuyNumber',
                    'category',
                    )     
    list_filter = (UserCategoryFilter, )  # 添加筛选器
    
    search_fields = ['TOKEN_ADDRESS', 'TOKEN_price',  ]

    
    # actions = (CreateTokenAddr, imputationAddr,your_action,'layerCreateToken','layerImputationToken')

    # actions = ('show_template_view','import_data','layerCreateToken','layerImputationToken','layerImputationTokenYmii',allTokenPrice,TokenBuySellZero,export_excel)
    actions = ('show_template_view','import_data','layerCreateToken','layerImputationToken','layerImputationTokenYmii',allTokenPrice,'TokenOtherPrice',TokenBuySellZero,export_excel)
 
    date_hierarchy = 'created_at'
    
    list_per_page = 15
    def taskLog(self, obj):
        
        parameter_str = '{} '.format(str(obj.QuantifyId) )
        btn_str = '<a class="btn btn-xs btn-danger" href="{}">' \
                  '<input name="查看明细"' \
                  'type="button" id="passButton" ' \
                  'title="passButton" value="查看明细">' \
                  '</a>'
                  
                  
        return format_html(btn_str, '/app1/showlog/{}/'.format(parameter_str))

    taskLog.short_description = '明细'


    def taskStop(self, obj): 
        parameter_str = 'id={} '.format(str(obj.id) )
        btn_str = '<a class="btn btn-xs btn-danger" href="{}">' \
                    '<input name="停止"' \
                    'type="button" id="passButton" ' \
                    'title="passButton" value="停止">' \
                    '</a>'
        return format_html(btn_str, '/app1/stopTask/?{}'.format(parameter_str))
    
    # form 部分
    fields =('TOKEN_ADDRESS',  'category', 'TOKEN_bak',             
              'BuyTOKEN_price', 'SellTOKEN_price',  'SellNumber', 'BuyNumber',
              'status',
             )

    readonly_fields = (  'TOKEN_ADDRESS', 'created_at',)
    

    taskStop.short_description = '停止'
    # 生成钱包
    def layerCreateToken(self, request, queryset):
        
        # reMsg=self.workClose(request, queryset)
        print('sssssss')
        reMsg='建立成功'        
        
        acct=request.user
        post = request.POST
        name=post.get('name')
        
        try:
            int_value = int(name)
            # print("Conversion successful:", int_value)
        except ValueError:
            # print("Conversion failed: The value is not convertible to int.")
            return JsonResponse(data={
                'status': 'success',
                'msg':  "请输入数字"
            })
            
        walletList=createWallet.createNewTokenWallet(int_value)
        # from cryptography.fernet import Fernet
        
        cipher_suite = Fernet('FRrgJ5KhZNgtdctTbQWV2-Xam9zZP-pdsUsLDDPg0pY=')

        for wallet in walletList:           
            text = str(wallet['privateKey'])
            cipher_text = cipher_suite.encrypt(text.encode('utf-8')).decode()            
            T_TokenAddr.objects.create(TOKEN_ADDRESS=wallet['address'],
                                       TOKEN_private=cipher_text,
                                       TOKEN_public=wallet['publicKey'],
                                       uid=acct,
                                       )
            T_TokenAddrBak.objects.create(TOKEN_ADDRESS=wallet['address'],
                                       TOKEN_private=cipher_text,
                                       TOKEN_public=wallet['publicKey'],
                                       uid=acct,
                                       )
      
        return JsonResponse(data={
                'status': 'success',
                'msg': reMsg
            })


    layerCreateToken.short_description = '生成钱包'
    layerCreateToken.type = 'success'
    layerCreateToken.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    layerCreateToken.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        # 'title': '弹出层输入框',
        # 提示信息
        'tips': '生成钱包',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [
            {
            # 这里的type 对应el-input的原生input属性，默认为input
            'type': 'input',
            # key 对应post参数中的key
            'key': 'name',
            # 显示的文本
            'label': '生成数量',
            # 为空校验，默认为False
            'require': True
        },       
                   ]
    }
    

     # 归集bnb
     # 如果你希望该视图仅对已登录用户可用，你还可以使用@login_required装饰器
    # @login_required
    def layerImputationToken(self, request, queryset):        
        # reMsg=self.workClose(request, queryset)       
        reMsg='归集成功:'                
        acct=request.user
        post = request.POST
        # 得到归集地址
        ImputationToken=post.get('name')        
        ImputationList=GetBnbWeb.GetBnb(request,ImputationToken)   

        reMsg=reMsg+str(len(ImputationList))
        
        # return render(request,"info.html",{"form":form})

      

        # return render(request, 'tokenlist.html', {'data_list': data_list})
        # return render(request, 'tokenlist.html')


        return JsonResponse(data={
                'status': 'success',
                'msg': str(reMsg)
                
            })


    layerImputationToken.short_description = '归集bnb'
    layerImputationToken.type = 'success'
    layerImputationToken.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    layerImputationToken.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        # 'title': '弹出层输入框',
        # 提示信息
        'tips': '归集bnb',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [
            {
            # 这里的type 对应el-input的原生input属性，默认为input
            'type': 'input',
            # key 对应post参数中的key
            'key': 'name',
            # 显示的文本
            'label': '归集地址',
            # 为空校验，默认为False
            'require': True
        },     ]
    }

# 归集ymii  或其他代币
     # 如果你希望该视图仅对已登录用户可用，你还可以使用@login_required装饰器
    # @login_required
    def layerImputationTokenYmii(self, request, queryset):        
        # reMsg=self.workClose(request, queryset)       
        reMsg='归集成功:'                
        acct=request.user
        post = request.POST
        # 得到归集地址
        ImputationToken=post.get('imputationToken') 
        tokenAddress=post.get('tokenAddress')         
        ImputationList=GetBnbWeb.GetYmii(request,ImputationToken,tokenAddress)   
        reMsg=reMsg+str(len(ImputationList)) 
      
        return JsonResponse(data={
                'status': 'success',
                'msg': reMsg
            })


    layerImputationTokenYmii.short_description = '归集代币(Ymii)'
    layerImputationTokenYmii.type = 'success'
    layerImputationTokenYmii.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    layerImputationTokenYmii.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        # 'title': '弹出层输入框',
        # 提示信息
        'tips': '归集代币',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [
            {
            # 这里的type 对应el-input的原生input属性，默认为input
            'type': 'input',
            # key 对应post参数中的key
            'key': 'imputationToken',
            # 显示的文本
            'label': '归集地址',
            # 为空校验，默认为False
            'require': True
        },
             {
            # 这里的type 对应el-input的原生input属性，默认为input
            'type': 'input',
            # key 对应post参数中的key
            'key': 'tokenAddress',
            # 显示的文本
            'label': '代币地址',    
             # 设置默认值
            'value': '0x768a62a22b187EB350637e720ebC552D905c0331'    ,    
            # 为空校验，默认为False
            'require': True
        },
            
            ]
    }
    
    
    
    def get_layer(self, request):
        """返回一个包含动态选项的layer字典。"""
        return {
            'tips': '导入数据',
            'confirm_button': '确认导入',
            'cancel_button': '取消',
            'width': '50%',
            'labelWidth': "100px",
            'params': [
                {
                    'type': 'select',
                    'key': 'category',
                    'label': '选择分类',
                    'require': True,
                    'options': [(category.id, category.name) for category in CategoryToken.objects.filter(uid=request.user)],
                },
                {
                    'type': 'textarea',
                    'key': 'data',
                    'label': '数据',
                    'require': True,
                    'placeholder': '每行一个数据'
                }
            ]
        }
        
   

    def show_template_view(modeladmin, request, queryset):
        print('dk')        
        # if not queryset:
        # 执行您希望在没有对象被选中时执行的代码
        # ...
        # return
    # 我们会在后面创建这个视图函数    
        return HttpResponseRedirect(reverse('custom_template_view'))
    
    show_template_view.short_description = '导入钱包'
    show_template_view.type = 'success'
    show_template_view.icon = 'el-icon-s-promotion'


    def import_data(self, request, queryset):      
        
        # self.layer =self.import_data_layer_info
  
        # reMsg=self.workClose(request, queryset)       
        reMsg='导入成功:'                
        acct=request.user
        # post = request.POST
        # # 得到归集地址
        # ImputationToken=post.get('category') 
        # tokenAddress=post.get('data')         
      
        
        
            # 从POST请求中获取数据
        category_id = request.POST.get('category') 
        data_text = request.POST.get('data')  

        # 分割数据，按行处理
        data_lines = data_text.strip().splitlines()

        # 使用数据库事务来确保所有操作都成功完成
        with transaction.atomic():
            for line in data_lines:
                # 假设每行是一个TOKEN_private
                token_private_data = line.strip()
                
                T_TokenAddr.objects.create(
                    # TOKEN_ADDRESS=
                    # wallList=createWallet.dbWallet(token_private_data),
                    TOKEN_private=createWallet.dbWallet(token_private_data),
                    category_id=category_id,
                    uid=acct
                )
                T_TokenAddrBak.objects.create(
                    # TOKEN_ADDRESS=wallet['address'],
                    TOKEN_private=createWallet.dbWallet(token_private_data),
                    category_id=category_id,
                    uid=acct
                                )
        # return HttpResponseRedirect('/admin/app1/t_tokenaddr/')  # 重定向到列表页面
      
        return JsonResponse(data={
                'status': 'success',
                'msg': reMsg
            })


    # abc=None
    
    def get_category_options(self, request):       
        # categories = CategoryToken.objects.all()
        categories =CategoryToken.objects.filter(uid=request.user)
        res= [{'key': cat.id, 'label': cat.name} for cat in categories]
        return res

  

        
    import_data.short_description = '导入token'
    import_data.type = 'success'
    import_data.icon = 'el-icon-s-promotion'
    
    # def changelist_view(self, request, extra_context=None):
    #     # 动态设置layer的options
    #     self.import_data.layer['params'][0]['options'] = self.get_category_options()
        
    #     # 调用父类的changelist_view方法
    #     return super().changelist_view(request, extra_context)
    
    import_data.layer = {
    'tips': '导入数据',
    'confirm_button': '确认导入',
    'cancel_button': '取消',
    'width': '50%',
    'labelWidth': "100px",
    'params': [
        {
            'type': 'select',
            'key': 'category', 
            'label': '选择分类',
            'require': False,
            'value': '0',
            # 根据您的模型获取分类选项
            'options':  [{'key':' cat.id', 'label': 'cat.name'} ] ,
        },
        {
            'type': 'textarea',
            'key': 'data',
            'label': '数据',
            'require': True,
            'placeholder': '每行一个数据'
        }
    ]
}
    def TokenOtherPrice(self, request, queryset):   
        reMsg='获取成功:'                
        tokenAddress=request.POST.get('tokenAddress')         
        tokenAddress=request.POST.get('tokenAddress')         

        # ImputationList=GetBnbWeb.GetYmii(request,ImputationToken,tokenAddress)   
            
        print("ok")
        # T_TokenAddrList=T_TokenAddr.objects.filter(uid=acct.id,status=1) 
        token_private_list = [obj  for obj in queryset]

        ImputationList=GetBnbWeb.GetOtherPrice(request,tokenAddress,token_private_list)   
        print (str(ImputationList))

        for item in ImputationList:
            token_addr = item[1].split(":")[1]  # 分割字符串取得token地址
            price = item[3]

            price=convert_integer_str_to_18_decimal_places(str(price))

            print(convert_integer_str_to_18_decimal_places(price))  # 输出：0.000000000000001234


            # 查找并更新
            token_obj = T_TokenAddr.objects.filter(TOKEN_ADDRESS=token_addr).first()
            if token_obj:
                token_obj.TOKEN_otherPrice = str(price)
                token_obj.save()

      
        return JsonResponse(data={
                'status': 'success',
                'msg': reMsg
            })


    TokenOtherPrice.short_description = '代币价格'
    TokenOtherPrice.type = 'success'
    TokenOtherPrice.icon = 'el-icon-s-promotion'

    # 指定一个输入参数，应该是一个数组

    # 指定为弹出层，这个参数最关键
    TokenOtherPrice.layer = {
        # 弹出层中的输入框配置

        # 这里指定对话框的标题
        # 'title': '弹出层输入框',
        # 提示信息
        'tips': '归集代币',
        # 确认按钮显示文本
        'confirm_button': '确认提交',
        # 取消按钮显示文本
        'cancel_button': '取消',

        # 弹出层对话框的宽度，默认50%
        'width': '40%',

        # 表单中 label的宽度，对应element-ui的 label-width，默认80px
        'labelWidth': "80px",
        'params': [
            {
            # 这里的type 对应el-input的原生input属性，默认为input
            'type': 'input',
            # key 对应post参数中的key
            'key': 'tokenAddress',
            # 显示的文本
            'label': '代币地址',    
             # 设置默认值
            'value': '0x768a62a22b187EB350637e720ebC552D905c0331'    ,    
            # 为空校验，默认为False
            'require': True
        },           
            {

            'type': 'select',

            'key': 'type',

            'label': '类型',

            'width': '200px',

            # size对应elementui的size，取值为：medium  small  mini

            'size': 'small',

            # value字段可以指定默认值

            'value': '0',

            'options': [{

                'key': '0',

                'label': '收入'

            }, {

                'key': '1',

                'label': '支出'

            }]

        },
            
            
            
            ]
    }
    
    # def get_dynamic_options_for_select(self):
    #     # 返回一个列表，包含下拉框的动态选项
    #     # 每个选项可以是一个字典，包含'label'和'value'键        
    #     categories = CategoryToken.objects.all()
    #     res= [{'key': cat.id, 'label': cat.name} for cat in categories]
    #     return res
    
    def changelist_view(self, request, extra_context=None):
        user_id = request.user.id
        # ... 这里做你需要的处理
        # ... 设置你需要的内容，例如调整layer的配置
        categoryList=self.get_category_options(request)
        self.TokenOtherPrice.layer['params'][0]['value'] = '123'
        self.import_data.layer['params'][0]['options'] = categoryList
        self.import_data.layer['params'][0]['value']=categoryList[0]['key']
        
        
        self.TokenOtherPrice.layer['params'][1]['options'] = categoryList
        self.TokenOtherPrice.layer['params'][1]['value']=categoryList[0]['key']       
         


        # 确保你调用了父类的changelist_view以正常显示列表
        return super().changelist_view(request, extra_context=extra_context)

        
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # self.TokenOtherPrice.layer['params'][0]['value'] = '123'
        # self.import_data.layer['params'][0]['options'] = self.get_dynamic_options_for_select()
        # self.TokenOtherPrice.layer['params'][1]['options'] = self.get_dynamic_options_for_select()
    








@admin.register(T_trade) 
class T_tradeAdmin(AjaxAdmin): 
    #    列表部分
    def get_queryset(self, request):
        queryset = super().get_queryset(request)
        accUser = request.user 
        # 在此处添加任何过滤逻辑
        queryset = queryset.filter(uid=accUser,status=1)
        return queryset
    
    # def has_add_permission(self, request):
    #     return False
    # def has_delete_permission(self, request, obj=None):
    #     return False
    
    list_display = ('TOKEN_ADDRESS', 'tradeToken', 'tradeStatus', 'Price' ,'tradeAmount', 'tradeHash','created_at' ) 
 
    date_hierarchy = 'created_at'
    
    list_per_page = 15
    # actions = [perform_check_all ]
    fieldsets = [('hash',{'fields':[('tradeHash','Price'),
                    'tradeAmount',   
                            ]}),
                (u'---',{'fields':['TOKEN_ADDRESS', 
                    'tradeToken','taskRemark',  
                    ]})]

    readonly_fields = ('tradeHash','Price','tradeAmount', 'TOKEN_ADDRESS', 
                    'tradeToken','taskRemark' ,) 




  

from django import forms


class CategoryAdmin(admin.ModelAdmin):
    form = CategoryForm
    
    def get_queryset(self, request):
        # 默认情况下只显示当前用户的分类
        qs = super(CategoryAdmin, self).get_queryset(request)
        # if request.user.is_superuser:
        #     return qs
        return qs.filter(uid=request.user)
    
    def get_form(self, request, obj=None, **kwargs):
        FormClass = super(CategoryAdmin, self).get_form(request, obj, **kwargs)
        return type('CategoryFormWrapper', (FormClass,), {'__init__': lambda self, *args, **kwargs: FormClass.__init__(self, *args, user=request.user, **kwargs)})


    def save_model(self, request, obj, form, change):
        if not change:
            obj.uid = request.user
        super().save_model(request, obj, form, change)

admin.site.register(CategoryToken, CategoryAdmin)
